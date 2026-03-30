import csv
import io
from datetime import date, datetime, timezone
from typing import Any

from fastapi import HTTPException
from openpyxl import load_workbook
from sqlalchemy import delete, select
from sqlalchemy.orm import Session

from app.models.import_job import ImportDataset, ImportJob, ImportJobError, ImportJobStatus
from app.models.inventory_override import InventoryOverride
from app.models.request_code_rule import RequestCodeRule
from app.models.reservation import Reservation
from app.models.room import Room

REQUIRED_COLUMNS: dict[ImportDataset, set[str]] = {
    ImportDataset.ROOM_MASTER: {"room_number", "room_type", "bed_type", "floor"},
    ImportDataset.REQUEST_CODE_RULES: {"code", "description", "default_weight"},
    ImportDataset.RESERVATIONS: {
        "external_id",
        "guest_name",
        "arrival_date",
        "departure_date",
        "requested_room_type",
    },
    ImportDataset.INVENTORY_OVERRIDES: {"room_number", "override_date", "capacity_delta", "status", "reason"},
}


def parse_file(filename: str, content: bytes) -> list[dict[str, Any]]:
    lower_name = filename.lower()
    if lower_name.endswith(".csv"):
        text_stream = io.StringIO(content.decode("utf-8-sig"))
        return [dict(row) for row in csv.DictReader(text_stream)]
    if lower_name.endswith(".xlsx"):
        workbook = load_workbook(io.BytesIO(content), data_only=True)
        sheet = workbook.active
        headers = [str(c.value).strip() if c.value is not None else "" for c in next(sheet.iter_rows(min_row=1, max_row=1))]
        rows: list[dict[str, Any]] = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            values = ["" if value is None else value for value in row]
            rows.append({headers[idx]: values[idx] if idx < len(values) else "" for idx in range(len(headers))})
        return rows
    raise HTTPException(status_code=400, detail="Only CSV and XLSX files are supported")


def _required_column_errors(dataset: ImportDataset, rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    if not rows:
        return [{"row_number": 0, "column_name": "file", "message": "File contains no data rows"}]
    incoming_columns = {key.strip() for key in rows[0].keys() if key}
    missing = REQUIRED_COLUMNS[dataset] - incoming_columns
    return [
        {"row_number": 0, "column_name": column, "message": "Missing required column"}
        for column in sorted(missing)
    ]


def _parse_date(value: Any) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    try:
        return datetime.fromisoformat(str(value)).date()
    except ValueError:
        return None




def _to_json_safe(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    safe_rows: list[dict[str, Any]] = []
    for row in rows:
        safe_row: dict[str, Any] = {}
        for key, value in row.items():
            if isinstance(value, (date, datetime)):
                safe_row[key] = value.isoformat()
            else:
                safe_row[key] = value
        safe_rows.append(safe_row)
    return safe_rows


def validate_rows(dataset: ImportDataset, rows: list[dict[str, Any]], db: Session) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    normalized_rows: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = _required_column_errors(dataset, rows)

    if errors:
        return normalized_rows, errors

    for idx, row in enumerate(rows, start=1):
        normalized: dict[str, Any] = {k.strip(): row.get(k) for k in row.keys()}

        def add_error(column: str, message: str) -> None:
            errors.append({"row_number": idx, "column_name": column, "message": message})

        if dataset is ImportDataset.ROOM_MASTER:
            if not str(normalized.get("room_number", "")).strip():
                add_error("room_number", "Required")
            if not str(normalized.get("room_type", "")).strip():
                add_error("room_type", "Required")
            if not str(normalized.get("bed_type", "")).strip():
                add_error("bed_type", "Required")
            try:
                normalized["floor"] = int(normalized.get("floor"))
            except (TypeError, ValueError):
                add_error("floor", "Must be an integer")
            for flag in ["is_hardblocked", "is_accessible", "is_near_elevator", "is_club_floor"]:
                normalized[flag] = str(normalized.get(flag, "false")).strip().lower() in {"1", "true", "yes", "y"}

        elif dataset is ImportDataset.REQUEST_CODE_RULES:
            if not str(normalized.get("code", "")).strip():
                add_error("code", "Required")
            if not str(normalized.get("description", "")).strip():
                add_error("description", "Required")
            try:
                normalized["default_weight"] = float(normalized.get("default_weight"))
            except (TypeError, ValueError):
                add_error("default_weight", "Must be numeric")

        elif dataset is ImportDataset.RESERVATIONS:
            for col in ["external_id", "guest_name", "requested_room_type", "requested_bed_type"]:
                if not str(normalized.get(col, "")).strip():
                    add_error(col, "Required")
            arrival = _parse_date(normalized.get("arrival_date"))
            departure = _parse_date(normalized.get("departure_date"))
            if arrival is None:
                add_error("arrival_date", "Must be ISO date YYYY-MM-DD")
            if departure is None:
                add_error("departure_date", "Must be ISO date YYYY-MM-DD")
            if arrival and departure and departure <= arrival:
                add_error("departure_date", "Must be after arrival_date")
            normalized["arrival_date"] = arrival
            normalized["departure_date"] = departure
            normalized["club_access_entitled"] = str(normalized.get("club_access_entitled", "false")).strip().lower() in {"1", "true", "yes", "y"}

        elif dataset is ImportDataset.INVENTORY_OVERRIDES:
            room_number = str(normalized.get("room_number", "")).strip()
            if not room_number:
                add_error("room_number", "Required")
            else:
                room_exists = db.scalar(select(Room.id).where(Room.room_number == room_number))
                if room_exists is None:
                    add_error("room_number", "Unknown room_number")
            override_date = _parse_date(normalized.get("override_date"))
            if override_date is None:
                add_error("override_date", "Must be ISO date YYYY-MM-DD")
            normalized["override_date"] = override_date
            try:
                normalized["capacity_delta"] = int(normalized.get("capacity_delta"))
            except (TypeError, ValueError):
                add_error("capacity_delta", "Must be an integer")
            if not str(normalized.get("status", "")).strip():
                add_error("status", "Required")
            if not str(normalized.get("reason", "")).strip():
                add_error("reason", "Required")

        normalized_rows.append(normalized)

    return normalized_rows, errors


def create_preview_job(dataset: ImportDataset, filename: str, rows: list[dict[str, Any]], db: Session) -> ImportJob:
    normalized_rows, errors = validate_rows(dataset, rows, db)
    invalid_row_numbers = {err["row_number"] for err in errors if err["row_number"] > 0}
    valid_rows = len([row for i, row in enumerate(normalized_rows, start=1) if i not in invalid_row_numbers])

    job = ImportJob(
        dataset=dataset,
        status=ImportJobStatus.PREVIEW_READY,
        filename=filename,
        total_rows=len(rows),
        valid_rows=valid_rows,
        invalid_rows=len(invalid_row_numbers),
        preview_rows_json=_to_json_safe(normalized_rows),
    )
    db.add(job)
    db.flush()

    for err in errors:
        db.add(
            ImportJobError(
                import_job_id=job.id,
                row_number=err["row_number"],
                column_name=err["column_name"],
                message=err["message"],
            )
        )

    db.commit()
    db.refresh(job)
    return job


def get_job_or_404(job_id: int, db: Session) -> ImportJob:
    job = db.get(ImportJob, job_id)
    if job is None:
        raise HTTPException(status_code=404, detail="Import job not found")
    return job


def get_job_errors(job_id: int, db: Session) -> list[ImportJobError]:
    return list(db.scalars(select(ImportJobError).where(ImportJobError.import_job_id == job_id).order_by(ImportJobError.row_number)))


def commit_job(job: ImportJob, db: Session) -> int:
    if job.status != ImportJobStatus.PREVIEW_READY:
        raise HTTPException(status_code=400, detail="Only preview_ready jobs can be committed")

    errors = get_job_errors(job.id, db)
    if any(err.row_number > 0 for err in errors):
        raise HTTPException(status_code=400, detail="Cannot commit import job with row validation errors")

    applied_rows = 0

    if job.dataset is ImportDataset.ROOM_MASTER:
        for row in job.preview_rows_json:
            room_number = str(row["room_number"]).strip()
            existing = db.scalar(select(Room).where(Room.room_number == room_number))
            if existing is None:
                db.add(
                    Room(
                        room_number=room_number,
                        room_type=str(row["room_type"]).strip(),
                        bed_type=str(row["bed_type"]).strip(),
                        floor=int(row["floor"]),
                        is_active=True,
                        is_hardblocked=bool(row.get("is_hardblocked", False)),
                        is_accessible=bool(row.get("is_accessible", False)),
                        is_near_elevator=bool(row.get("is_near_elevator", False)),
                        is_club_floor=bool(row.get("is_club_floor", False)),
                    )
                )
            else:
                existing.room_type = str(row["room_type"]).strip()
                existing.bed_type = str(row["bed_type"]).strip()
                existing.floor = int(row["floor"])
                existing.is_hardblocked = bool(row.get("is_hardblocked", False))
                existing.is_accessible = bool(row.get("is_accessible", False))
                existing.is_near_elevator = bool(row.get("is_near_elevator", False))
                existing.is_club_floor = bool(row.get("is_club_floor", False))
            applied_rows += 1

    elif job.dataset is ImportDataset.REQUEST_CODE_RULES:
        for row in job.preview_rows_json:
            code = str(row["code"]).strip()
            existing = db.scalar(select(RequestCodeRule).where(RequestCodeRule.code == code))
            if existing is None:
                db.add(
                    RequestCodeRule(
                        code=code,
                        description=str(row["description"]).strip(),
                        default_weight=float(row["default_weight"]),
                    )
                )
            else:
                existing.description = str(row["description"]).strip()
                existing.default_weight = float(row["default_weight"])
            applied_rows += 1

    elif job.dataset is ImportDataset.RESERVATIONS:
        for row in job.preview_rows_json:
            external_id = str(row["external_id"]).strip()
            existing = db.scalar(select(Reservation).where(Reservation.external_id == external_id))
            payload = {
                "guest_name": str(row["guest_name"]).strip(),
                "arrival_date": _parse_date(row["arrival_date"]),
                "departure_date": _parse_date(row["departure_date"]),
                "requested_room_type": str(row["requested_room_type"]).strip(),
                "requested_bed_type": str(row["requested_bed_type"]).strip(),
                "club_access_entitled": bool(row.get("club_access_entitled", False)),
                "status": "booked",
            }
            if existing is None:
                db.add(Reservation(external_id=external_id, **payload))
            else:
                for key, value in payload.items():
                    setattr(existing, key, value)
            applied_rows += 1

    elif job.dataset is ImportDataset.INVENTORY_OVERRIDES:
        for row in job.preview_rows_json:
            room_number = str(row["room_number"]).strip()
            room = db.scalar(select(Room).where(Room.room_number == room_number))
            if room is None:
                raise HTTPException(status_code=400, detail=f"Room '{room_number}' no longer exists")
            db.add(
                InventoryOverride(
                    room_id=room.id,
                    override_date=_parse_date(row["override_date"]),
                    capacity_delta=int(row["capacity_delta"]),
                    status=str(row["status"]).strip(),
                    reason=str(row["reason"]).strip(),
                )
            )
            applied_rows += 1

    else:
        raise HTTPException(status_code=400, detail="Unsupported dataset")

    job.status = ImportJobStatus.COMMITTED
    job.committed_at = datetime.now(timezone.utc)
    db.execute(delete(ImportJobError).where(ImportJobError.import_job_id == job.id, ImportJobError.row_number == 0))
    db.commit()
    db.refresh(job)
    return applied_rows
